import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

def generate_schedule_table(start_date_str, end_date_str, extra_rows=10, output_file="schedule_table.xlsx"):
    # 날짜 파싱
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

    # 날짜 범위 생성 (토요일, 일요일 제외)
    date_range = pd.date_range(start=start_date, end=end_date, freq='B')

    # 엑셀 워크북과 시트 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "프로젝트 예상 일정"

    # 셀 스타일 설정
    header_fill = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")  # 월 헤더 색상
    week_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # 주 헤더 색상
    day_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # 일 헤더 색상
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 첫 줄 빈 줄
    ws.append([])

    # 프로젝트 제목 추가 (좌측 정렬, 셀 병합 없음)
    ws["B2"] = "프로젝트 일정 계획표"
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"].font = Font(size=16, bold=True)

    # 한 줄 띄우기
    ws.append([])

    # 카테고리, 테스크 명 컬럼 생성 (아래 두 줄과 병합)
    ws.merge_cells("B4:B6")
    ws["B4"] = "카테고리"
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B4"].fill = header_fill
    ws["B4"].border = thin_border

    ws.merge_cells("C4:C6")
    ws["C4"] = "테스크 명"
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C4"].fill = header_fill
    ws["C4"].border = thin_border

    # "월" 정보를 생성하고 병합 없이 배치
    month_positions = {}
    for date in date_range:
        month = date.strftime("%Y년 %m월")
        if month not in month_positions:
            month_positions[month] = []
        month_positions[month].append(date)

    col_start = 4  # 실제 데이터가 들어가는 시작 열 (카테고리와 테스크 명이 2~3열)
    for month, dates in month_positions.items():
        col_end = col_start + len(dates) - 1
        ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)
        ws.cell(row=4, column=col_start).value = month
        ws.cell(row=4, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=col_start).fill = header_fill
        col_start = col_end + 1

    # "주" 정보 생성 (첫 주를 1로 설정)
    col_start = 4
    week_number = 1
    current_week_start = None

    for date in date_range:
        if date.weekday() == 0 or current_week_start is None:
            if current_week_start is not None:
                col_end = col_start - 1
                ws.merge_cells(start_row=5, start_column=current_week_start, end_row=5, end_column=col_end)
                ws.cell(row=5, column=current_week_start).value = f"W{week_number}"
                ws.cell(row=5, column=current_week_start).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=5, column=current_week_start).fill = week_fill
                week_number += 1

            current_week_start = col_start

        # "일" 헤더 생성
        ws.cell(row=6, column=col_start).value = date.day
        ws.cell(row=6, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=col_start).fill = day_fill
        col_start += 1

    # 마지막 주 처리
    ws.merge_cells(start_row=5, start_column=current_week_start, end_row=5, end_column=col_start - 1)
    ws.cell(row=5, column=current_week_start).value = f"W{week_number}"
    ws.cell(row=5, column=current_week_start).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=current_week_start).fill = week_fill

    # 셀 너비 설정 (카테고리와 테스크 셀 넓이를 넓게 설정)
    ws.column_dimensions['A'].width = 5  # 카테고리
    ws.column_dimensions['B'].width = 20  # 카테고리
    ws.column_dimensions['C'].width = 30  # 테스크 명
    for col in range(4, col_start):
        ws.column_dimensions[get_column_letter(col)].width = 3.3

    # 전체 테두리 얇게 설정 (B4:AZ16)
    for row in range(4, extra_rows):  # 범위를 빈 줄까지 확장
        for col in range(2, col_start):
            ws.cell(row=row, column=col).border = thin_border



    # 월화수목금 사이의 테두리를 없애는 코드
    for col in range(4, col_start):  # 월요일부터 금요일까지의 열 범위 (열 4부터 시작)
        if (col - 3) % 5 == 1:  # 월요일
            ws.cell(row=6, column=col).border = Border(
                left=thin_border.left,
                right=Side(border_style=None),  # 오른쪽 테두리 제거 (화요일과의 경계)
                top=thin_border.top,
                bottom=thin_border.bottom
            )
        elif (col - 3) % 5 in [2, 3, 4]:  # 화요일, 수요일, 목요일
            ws.cell(row=6, column=col).border = Border(
                left=Side(border_style=None),  # 왼쪽 테두리 제거
                right=Side(border_style=None),  # 오른쪽 테두리 제거
                top=thin_border.top,
                bottom=thin_border.bottom
            )
        elif (col - 3) % 5 == 0:  # 금요일
            ws.cell(row=6, column=col).border = Border(
                left=Side(border_style=None),  # 왼쪽 테두리 제거 (목요일과의 경계)
                right=thin_border.right,  # 오른쪽 테두리 유지
                top=thin_border.top,
                bottom=thin_border.bottom
            )



    # 주별 구분을 위한 굵은 테두리 설정
    for row in range(7, extra_rows):  # 데이터 행 범위
        for col in range(4, col_start, 5):  # 월요일에 해당하는 열마다 굵은 테두리
            ws.cell(row=row, column=col).border = thick_border


    # 엑셀 파일로 저장
    wb.save(output_file)
    print(f"Schedule table saved to {output_file}")

# 함수 실행 예시 (빈 줄 추가 가능)
generate_schedule_table("2024-08-26", "2025-02-28", extra_rows=20)
